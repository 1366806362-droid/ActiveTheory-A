#!/usr/bin/env python3
"""Convert the locked 2026-07-29 GEO/5A fusion inputs to Dashboard schema 1.0.0.

This is an offline, deterministic converter. It never modifies its ZIP inputs and
never copies source workbooks into the repository.
"""

from __future__ import annotations

import argparse
import collections
import datetime as dt
import hashlib
import io
import json
import math
import os
import re
import tempfile
import unicodedata
import zipfile
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_MAPPING = Path(__file__).with_name("geo_real_data_mapping_20260729.json")
MANIFEST_PATH = ROOT / "public" / "data" / "geo-dashboard" / "manifest.json"
FUSION_REPORT = "geo_5a_analysis_report_2026-07-29.json"
QUALITY_REPORT = "data_quality_validation_2026-07-29.json"
ACCESS_REPORT = "data_access_validation.json"
FUSION_SUMMARY = "geo_5a_analysis_summary_2026-07-29.xlsx"
RULE_TABLE_NAME = "GEO_Codex分析规则总表_V1.md"
RULE_CHANGELOG_NAME = "GEO_Codex规则变更日志_V1.md"


class ConversionError(RuntimeError):
    """Raised when a blocking source or reconciliation check fails."""


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_id(prefix: str, *parts: Any, length: int = 20) -> str:
    text = "\x1f".join("" if part is None else str(part) for part in parts)
    return f"{prefix}_{sha256_bytes(text.encode('utf-8'))[:length]}"


def normalize_question(value: Any) -> str:
    text = unicodedata.normalize("NFC", str(value or ""))
    return re.sub(r"\s+", " ", text.replace("\r\n", "\n").replace("\r", "\n")).strip()


def date_string(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    match = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    return match.group(1) if match else text


def number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value if math.isfinite(value) else None
    try:
        parsed = float(str(value).replace(",", "").rstrip("%"))
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed):
        return None
    return int(parsed) if parsed.is_integer() else parsed


def percent(value: Any) -> float | None:
    parsed = number(value)
    if parsed is None:
        return None
    return round(float(parsed) * 100 if abs(float(parsed)) <= 1 else float(parsed), 8)


def split_values(value: Any) -> list[str]:
    if value is None:
        return []
    return sorted({part.strip() for part in re.split(r"[,，、;；]", str(value)) if part.strip()})


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def json_bytes(value: Any) -> bytes:
    text = json.dumps(
        json_safe(value),
        ensure_ascii=False,
        indent=2,
        allow_nan=False,
    ) + "\n"
    return text.encode("utf-8")


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(json_bytes(value))


def source_file_info(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path.resolve()),
        "fileName": path.name,
        "sizeBytes": stat.st_size,
        "sha256": sha256_file(path),
        "lastModified": dt.datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
    }


def rule_file_info(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path.resolve()),
        "fileName": path.name,
        "sizeBytes": stat.st_size,
        "sha256": sha256_file(path),
        "lastModified": dt.datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
    }


def extract_rule_counts(text: str) -> dict[str, int]:
    headings = list(re.finditer(r"(?m)^###\s+(R-[A-Z0-9-]+)\uff5c", text))
    counts = {"effective": 0, "deprecated": 0, "pendingConfirmation": 0, "total": len(headings)}
    for index, match in enumerate(headings):
        end = headings[index + 1].start() if index + 1 < len(headings) else len(text)
        block = text[match.start():end]
        status_match = re.search(r"\*\*(?:版本/状态|当前状态)\*\*\uff1a([^\n]+)", block)
        status = status_match.group(1).strip() if status_match else ""
        if "已废弃" in status:
            counts["deprecated"] += 1
        elif "待确认" in status and "V2已确认" not in status:
            counts["pendingConfirmation"] += 1
        else:
            counts["effective"] += 1
    return counts


def answer_validity(platform: str, question: str, answer: str) -> tuple[bool, str | None]:
    if not platform:
        return False, "missing_platform"
    if not question:
        return False, "missing_question"
    if not answer:
        return False, "empty_answer"
    if normalize_question(answer) == normalize_question(question):
        return False, "answer_equals_question"
    if re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", answer.strip()):
        return False, "date_only_answer"
    return True, None


def safe_extract(archive: Path, destination: Path) -> list[str]:
    destination = destination.resolve()
    members: list[str] = []
    with zipfile.ZipFile(archive) as zipped:
        for info in zipped.infolist():
            target = (destination / info.filename).resolve()
            if destination not in target.parents and target != destination:
                raise ConversionError(f"ZIP_PATH_TRAVERSAL: {archive.name}!{info.filename}")
            zipped.extract(info, destination)
            members.append(info.filename)
    return members


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        records = []
        for source_row, values in enumerate(rows, start=2):
            if not any(value is not None for value in values):
                continue
            record = dict(zip(headers, values))
            record["__row__"] = source_row
            records.append(record)
        return headers, records
    finally:
        workbook.close()


def inventory_workbooks(root: Path) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*.xlsx"), key=lambda item: str(item).casefold()):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                first = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [str(value).strip() if value is not None else "" for value in first]
                entries.append({
                    "path": path,
                    "relativePath": path.relative_to(root).as_posix(),
                    "sheet": worksheet.title,
                    "headers": headers,
                    "rowCount": max(worksheet.max_row - 1, 0),
                    "columnCount": worksheet.max_column,
                })
        finally:
            workbook.close()
    return entries


def identify_table(entries: list[dict[str, Any]], required: Iterable[str], label: str) -> dict[str, Any]:
    required_set = set(required)
    matches = [entry for entry in entries if required_set.issubset(set(entry["headers"]))]
    if len(matches) != 1:
        raise ConversionError(f"TABLE_IDENTIFICATION_{label.upper()}: expected 1, found {len(matches)}")
    entry = matches[0]
    headers, rows = read_sheet(entry["path"], entry["sheet"])
    return {**entry, "headers": headers, "rows": rows}


def read_json_by_name(root: Path, file_name: str) -> dict[str, Any]:
    matches = list(root.rglob(file_name))
    if len(matches) != 1:
        raise ConversionError(f"SOURCE_JSON_{file_name}: expected 1, found {len(matches)}")
    return json.loads(matches[0].read_text(encoding="utf-8"))


def identify_fusion_workbook(entries: list[dict[str, Any]]) -> Path:
    grouped: dict[Path, set[str]] = collections.defaultdict(set)
    for entry in entries:
        grouped[entry["path"]].add(entry["sheet"])
    required = {"summary", "geo_semantic", "citation_classification_detail", "geo_keywords_expanded", "opportunity_all"}
    matches = [path for path, sheets in grouped.items() if required.issubset(sheets)]
    if len(matches) != 1:
        raise ConversionError(f"FUSION_WORKBOOK_IDENTIFICATION: expected 1, found {len(matches)}")
    return matches[0]


def table_from_fusion(path: Path, sheet: str) -> list[dict[str, Any]]:
    return read_sheet(path, sheet)[1]


def platform_id(value: Any, aliases: dict[str, str], warnings: list[dict[str, Any]]) -> str:
    raw = str(value or "").strip()
    if raw.casefold() in {"all", "doubao", "deepseek", "kimi", "qwen"}:
        return raw.casefold()
    mapped = aliases.get(raw) or aliases.get(raw.casefold())
    if mapped:
        return mapped
    safe = re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode().casefold()).strip("-")
    safe = safe or f"unknown-{sha256_bytes(raw.encode('utf-8'))[:10]}"
    warnings.append({
        "code": "UNKNOWN_PLATFORM",
        "message": f"未知平台“{raw}”已保留为安全ID“{safe}”。",
        "platform": raw,
    })
    return safe


def role_level(role: str) -> str:
    return {
        "主推荐": "primary",
        "次推荐": "secondary",
        "软植入": "softPlacement",
        "多品牌混合/歧义归属": "weak",
        "无品牌产品推荐": "weak",
        "产品错归竞品": "noRecommendation",
        "未提及": "noRecommendation",
    }.get(role, "noRecommendation")


def source_type(value: Any) -> str:
    return {
        "官方来源": "official",
        "第三方来源": "thirdParty",
        "社区来源": "community",
    }.get(str(value or "").strip(), "unknown")


def content_type(value: Any) -> str:
    return {
        "新闻媒体": "media",
        "论坛社区": "communityPost",
        "测评站": "review",
        "百科平台": "other",
        "内容站": "other",
        "电商平台": "other",
        "视频平台": "other",
        "其他": "other",
    }.get(str(value or "").strip(), "other")


def build_answer_records(
    raw_rows: list[dict[str, Any]],
    semantic_rows: list[dict[str, Any]],
    target_date: str,
    aliases: dict[str, str],
    warnings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    semantics: dict[tuple[str, str], collections.deque[dict[str, Any]]] = collections.defaultdict(collections.deque)
    for row in semantic_rows:
        semantics[(str(row.get("AI平台") or "").strip(), normalize_question(row.get("关键词提问")))].append(row)
    occurrences: collections.Counter[tuple[str, str, str, str]] = collections.Counter()
    records: list[dict[str, Any]] = []
    for row in raw_rows:
        if date_string(row.get("日期")) != target_date:
            continue
        platform_raw = str(row.get("AI平台") or "").strip()
        question = normalize_question(row.get("关键词提问"))
        queue = semantics[(platform_raw, question)]
        if not queue:
            raise ConversionError(f"ANSWER_SEMANTIC_JOIN_MISSING: {platform_raw} | {question}")
        semantic = queue.popleft()
        pid = platform_id(platform_raw, aliases, warnings)
        question_id = stable_id("q", question, length=16)
        conversation = str(row.get("对话分享链接") or "").strip()
        key = (target_date, pid, question_id, conversation)
        occurrence = occurrences[key]
        occurrences[key] += 1
        role = str(semantic.get("brand_role") or "未提及").strip()
        brand_mentioned = role in {"主推荐", "次推荐", "软植入", "多品牌混合/歧义归属"}
        raw_answer = str(row.get("AI回答") or "").strip()
        is_valid, invalid_reason = answer_validity(platform_raw, question, raw_answer)
        first_character_position = number(semantic.get("品牌首次出现位置"))
        records.append({
            "recordId": stable_id("ans", *key, occurrence),
            "date": target_date,
            "platformId": pid,
            "questionId": question_id,
            "question": question,
            "answerType": semantic.get("semantic_content_type") or "unknown",
            "isValid": is_valid,
            "brandMentioned": brand_mentioned,
            "brandPosition": None,
            "isFirstRecommendation": None,
            "recommendationLevel": role_level(role),
            "mentionedBrands": ["杨掌柜"] if brand_mentioned else [],
            "mentionedProducts": [semantic.get("entity_name")] if semantic.get("entity_name") else [],
            "rawReference": {
                "sourceModule": "GEO清洗结果",
                "sourceSheet": "AI回答分析表",
                "sourceRow": row["__row__"],
            },
            "diagnostics": {
                "brandRole": role,
                "attributionStatus": semantic.get("attribution_status"),
                "attributedBrand": semantic.get("attributed_brand"),
                "firstCharacterPosition": first_character_position,
                "firstCharacterPositionMeaning": "text_character_position_not_recommendation_rank",
                "invalidReason": invalid_reason,
                "duplicateOccurrenceIndex": occurrence,
            },
        })
    leftovers = sum(len(queue) for queue in semantics.values())
    if leftovers:
        raise ConversionError(f"ANSWER_SEMANTIC_JOIN_UNUSED: {leftovers}")
    return sorted(records, key=lambda item: (item["platformId"], item["questionId"], item["recordId"]))


def build_citation_records(
    raw_rows: list[dict[str, Any]],
    classified_rows: list[dict[str, Any]],
    target_date: str,
    aliases: dict[str, str],
    warnings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    classified: dict[tuple[str, str], collections.deque[dict[str, Any]]] = collections.defaultdict(collections.deque)
    for row in classified_rows:
        classified[(str(row.get("AI平台") or "").strip(), str(row.get("来源网址") or "").strip())].append(row)
    occurrences: collections.Counter[tuple[str, str, str, str]] = collections.Counter()
    records: list[dict[str, Any]] = []
    for row in raw_rows:
        if date_string(row.get("日期")) != target_date:
            continue
        platform_raw = str(row.get("AI平台") or "").strip()
        source_url = str(row.get("来源网址") or "").strip()
        queue = classified[(platform_raw, source_url)]
        if not queue:
            raise ConversionError(f"CITATION_CLASSIFICATION_JOIN_MISSING: {platform_raw} | {source_url}")
        classification = queue.popleft()
        pid = platform_id(platform_raw, aliases, warnings)
        question = normalize_question(row.get("关键词提问"))
        question_id = stable_id("q", question, length=16)
        key = (target_date, pid, question_id, source_url)
        occurrence = occurrences[key]
        occurrences[key] += 1
        ownership = classification.get("source_ownership") or "其他"
        stype = source_type(classification.get("官方/第三方/社区"))
        ctype = content_type(classification.get("平台类型"))
        records.append({
            "citationId": stable_id("cit", *key, occurrence),
            "date": target_date,
            "platformId": pid,
            "questionId": question_id,
            "url": None,
            "domain": classification.get("来源域名") or (urlparse(source_url).hostname or "unknown"),
            "title": row.get("来源标题") or None,
            "sourceType": stype,
            "sourceOwnership": ownership,
            "contentType": ctype,
            "isRanking": False,
            "isReview": ctype == "review",
            "isOfficial": ownership == "目标品牌官方来源",
            "isIndexed": None,
            "qualityLevel": None,
            "abnormalReason": None,
            "rawReference": {
                "sourceModule": "GEO清洗结果",
                "sourceSheet": "AI引用分析表",
                "sourceRow": row["__row__"],
            },
            "diagnostics": {"duplicateOccurrenceIndex": occurrence},
        })
    leftovers = sum(len(queue) for queue in classified.values())
    if leftovers:
        raise ConversionError(f"CITATION_CLASSIFICATION_JOIN_UNUSED: {leftovers}")
    return sorted(records, key=lambda item: (item["platformId"], item["questionId"], item["domain"], item["citationId"]))


def build_keyword_records(
    expanded_rows: list[dict[str, Any]],
    target_date: str,
    aliases: dict[str, str],
    warnings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    records = []
    occurrences: collections.Counter[tuple[str, str, str, int]] = collections.Counter()
    for row in expanded_rows:
        source_keyword = normalize_question(row.get("来源关键词"))
        candidate = normalize_question(row.get("拓展关键词"))
        expansion_index = int(number(row.get("expansion_index")) or 0)
        key = (target_date, source_keyword, candidate, expansion_index)
        duplicate = occurrences[key]
        occurrences[key] += 1
        link = str(row.get("AI对话链接") or "")
        pid = "doubao" if "doubao.com" in link.casefold() else "all"
        legacy_tested = str(row.get("是否已测试") or "").strip() == "是"
        legacy_triggered = str(row.get("是否触发AI引用") or "").strip() == "是"
        priority = {"高": "high", "中": "medium", "低": "low"}.get(str(row.get("优先级") or "").strip(), "unknown")
        records.append({
            "keywordId": stable_id("kw", *key, duplicate),
            "keyword": candidate,
            "normalizedKeyword": normalize_question(candidate).casefold(),
            "sourceKeyword": source_keyword,
            "candidateKeyword": candidate,
            "platformId": platform_id(pid, aliases, warnings),
            "date": target_date,
            "commercialValue": None,
            "aiTriggerType": "unknown",
            "brandOpportunity": None,
            "optimizationDirection": row.get("拓展方式") or None,
            "sceneType": " / ".join(filter(None, [str(row.get("分类") or "").strip(), str(row.get("用户意图") or "").strip()])) or "unknown",
            "trend": "flat",
            "trendValue": None,
            "priority": priority,
            "status": "candidate-untested",
            "rawReference": {
                "sourceModule": "融合分析结果",
                "sourceSheet": "geo_keywords_expanded",
                "sourceRow": row["__row__"],
            },
            "diagnostics": {
                "duplicateOccurrenceIndex": duplicate,
                "expansionIndex": expansion_index,
                "legacyTestedFlag": legacy_tested,
                "legacyCitationTriggeredFlag": legacy_triggered,
                "legacyStatusDeprecated": True,
            },
        })
    return sorted(records, key=lambda item: (item["sourceKeyword"], item["candidateKeyword"], item["keywordId"]))


def make_reconciliation(mapping: dict[str, Any]) -> tuple[list[dict[str, Any]], Any]:
    rows: list[dict[str, Any]] = []

    def check(name: str, expected: Any, actual: Any, source: str, *, tolerance: float = 0, blocking: bool = True) -> None:
        if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
            difference: Any = round(float(actual) - float(expected), 10)
            passed = abs(difference) <= tolerance
        else:
            difference = None if actual == expected else {"expected": expected, "actual": actual}
            passed = actual == expected
        rows.append({
            "check": name,
            "expected": expected,
            "actual": actual,
            "passed": passed,
            "source": source,
            "difference": difference,
            "blockingLevel": "error" if blocking else "warning",
        })

    return rows, check


def safety_scan(dataset: dict[str, Any]) -> dict[str, Any]:
    serialized = json.dumps(dataset, ensure_ascii=False, allow_nan=False)
    violations: list[dict[str, Any]] = []
    patterns = {
        "windowsUserPath": r"[A-Za-z]:\\Users\\",
        "fileUrl": r"file://",
        "aiShareLink": r"https?://[^\s\"']*(?:/share/|doubao\.com/thread)",
        "nonFiniteNumber": r"(?<![A-Za-z0-9_])(?:NaN|Infinity|-Infinity)(?![A-Za-z0-9_])",
        "unescapedControlCharacter": r"[\x00-\x08\x0b\x0c\x0e-\x1f]",
    }
    for name, pattern in patterns.items():
        match = re.search(pattern, serialized, flags=re.IGNORECASE)
        if match:
            violations.append({"check": name, "sample": match.group(0)[:80]})

    long_answers: list[str] = []

    def inspect(value: Any, path: str = "$") -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                child = f"{path}.{key}"
                if "answer" in key.casefold() and isinstance(item, str) and len(item) > 500:
                    long_answers.append(child)
                inspect(item, child)
        elif isinstance(value, list):
            for index, item in enumerate(value):
                inspect(item, f"{path}[{index}]")
        elif isinstance(value, float) and not math.isfinite(value):
            violations.append({"check": "nonFiniteValue", "path": path})

    inspect(dataset)
    if long_answers:
        violations.append({"check": "answerFieldOver500Characters", "paths": long_answers})
    return {
        "status": "pass" if not violations else "fail",
        "checks": {name: not any(item["check"] == name for item in violations) for name in patterns},
        "answerFieldLengthCheck": not long_answers,
        "violations": violations,
    }


def field_mapping_markdown() -> str:
    return """# GEO Dashboard V1.6.2 字段映射（2026-07-29）

| Dashboard字段 | 权威来源 | 工作表/字段 | 处理方式 |
|---|---|---|---|
| overview.finalScore | 融合分析JSON | scores.final_score | 直接读取，不重算 |
| overview.geoStructureScore | 融合分析JSON | scores.geo_structural_quality_score | 直接读取 |
| overview.geoSemanticScore | 融合分析JSON | scores.geo_semantic_score | 直接读取 |
| overview.geoScore | 融合分析JSON | scores.geo_score | 直接读取 |
| overview.fiveAScore | 融合分析JSON | scores.a5_score | 直接读取 |
| overview.industryOpportunityScore | 融合分析JSON | scores.industry_opportunity_driver_score | 直接读取 |
| answer.records | GEO清洗Excel | AI回答分析表 | 仅保留目标日期；不公开完整回答与分享链接 |
| answer语义字段 | 融合汇总Excel | geo_semantic | AI平台＋问题＋稳定出现顺序连接 |
| answer.brandPosition | 无可靠推荐排名 | — | 保持null；字符位置仅存diagnostics |
| answer.questionCollectionCompleteness | 正式规则总表＋AI回答明细 | 固定18题×4平台；实际平台×问题组合 | 49/72=68.06%，输出23条缺失矩阵 |
| answer.collectedAnswerValidity | 规则变更日志＋AI回答明细 | 回答与问题相同的源行 | 48/49=97.96%；无效行排除V2事实指标 |
| answer.firstRecommendationRate | 正式规则总表 | 首位推荐定义待确认 | 保持null；主推荐、次推荐和品牌推荐率独立展示 |
| citation.records | GEO清洗Excel | AI引用分析表 | 逐行明细；公开包不保留完整URL |
| citation分类 | 融合汇总Excel | citation_classification_detail | AI平台＋URL＋稳定出现顺序连接 |
| citation.qualityRate | 无逐条优质引用口径 | — | 保持null；73.63仅存内容来源质量诊断 |
| citation.indexedRate | 无可靠收录状态 | — | 保持null/missing |
| keyword.records | 融合汇总Excel | geo_keywords_expanded | 39条AI拓展候选直接读取，不能视为已实测 |
| keyword.topKeywords | 融合结果 | opportunity_top10 | 按execution_priority_rank排序 |
| keyword.commercialValue | 源数据未提供 | — | 保持null，禁止从优先级换算 |
| keyword.brandOpportunity | GEO拓展词未提供 | — | 保持null；行业机会词可使用明确brand_match_score |
| dataHealth | 正式规则总表＋数据质量JSON＋清洗明细 | 固定问题面板、platform_coverage/input_processing | 三项独立；含49/72完整率、48/49有效率和48/72理论有效完整率 |
| metadata快照日期 | 正式规则总表＋规则变更日志 | 7月22日源快照 | 5A/品牌心智均保留2026-07-22，lag=7；不使用历史改写日期 |
| trends | 单日逐行明细 | — | 输出空数组；聚合趋势仅存diagnostics |
| alerts | 明确质量事实 | 数据质量JSON/融合JSON | 不生成推测性业务结论 |
| recommendations | 融合分析JSON | scores.execution_order | 直接读取既有执行顺序 |

确定性派生仅包括：标准平台ID、稳定SHA-256 ID、域名计数、问题集/平台集/竞品集版本、逐行统计核验。明确禁止反推评分权重、品牌排名、优质引用率、问题计划分母和关键词商业价值。
"""


def convert(args: argparse.Namespace) -> dict[str, Any]:
    geo_zip = Path(args.geo_zip).resolve()
    a5_zip = Path(args.a5_zip).resolve()
    fusion_zip = Path(args.fusion_zip).resolve()
    output = Path(args.output).resolve()
    diagnostics_dir = Path(args.diagnostics_dir).resolve()
    mapping = json.loads(Path(args.mapping).read_text(encoding="utf-8"))
    target_date = args.target_date
    aliases = mapping["platformAliases"]
    business_rules = mapping["businessRules"]
    warnings: list[dict[str, Any]] = []

    rule_table = Path(getattr(args, "rule_table", "") or "").resolve()
    rule_changelog = Path(getattr(args, "rule_changelog", "") or "").resolve()
    if not rule_table.is_file() or rule_table.name != RULE_TABLE_NAME:
        raise ConversionError(f"RULE_TABLE_MISSING: {RULE_TABLE_NAME}")
    if not rule_changelog.is_file() or rule_changelog.name != RULE_CHANGELOG_NAME:
        raise ConversionError(f"RULE_CHANGELOG_MISSING: {RULE_CHANGELOG_NAME}")
    rule_table_text = rule_table.read_text(encoding="utf-8")
    rule_changelog_text = rule_changelog.read_text(encoding="utf-8")
    rule_sources = {
        "ruleTable": {**rule_file_info(rule_table), "documentVersion": business_rules["ruleDocumentVersion"]},
        "ruleChangelog": {**rule_file_info(rule_changelog), "documentVersion": "GEO_Codex规则变更日志_V1"},
        "usedAt": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "ruleCounts": extract_rule_counts(rule_table_text),
    }

    sources = {
        "geo": source_file_info(geo_zip),
        "a5": source_file_info(a5_zip),
        "fusion": source_file_info(fusion_zip),
    }
    diagnostics_dir.mkdir(parents=True, exist_ok=True)
    write_json(diagnostics_dir / "source-inventory.json", sources)
    correction_dir = diagnostics_dir / "business-rule-correction"
    correction_dir.mkdir(parents=True, exist_ok=True)
    write_json(correction_dir / "rule-source-inventory.json", rule_sources)

    with tempfile.TemporaryDirectory(prefix="active-theory-geo-v16-") as temporary:
        temporary_root = Path(temporary)
        extracted: dict[str, dict[str, Any]] = {}
        for label, source in (("geo", geo_zip), ("a5", a5_zip), ("fusion", fusion_zip)):
            destination = temporary_root / label
            destination.mkdir()
            members = safe_extract(source, destination)
            extracted[label] = {"root": destination, "members": members}

        geo_entries = inventory_workbooks(extracted["geo"]["root"])
        a5_entries = inventory_workbooks(extracted["a5"]["root"])
        fusion_entries = inventory_workbooks(extracted["fusion"]["root"])
        signatures = mapping["sheetSignatures"]
        ai_answer = identify_table(geo_entries, signatures["aiAnswer"], "ai_answer")
        ai_citation = identify_table(geo_entries, signatures["aiCitation"], "ai_citation")
        geo_keyword = identify_table(geo_entries, signatures["geoKeyword"], "geo_keyword")
        a5_asset = identify_table(a5_entries, signatures["a5Asset"], "a5_asset")
        a5_flow = identify_table(a5_entries, signatures["a5Flow"], "a5_flow")
        audience = identify_table(a5_entries, signatures["audience"], "audience")
        brand_mind = identify_table(a5_entries, signatures["brandMind"], "brand_mind")
        industry_opportunity = identify_table(a5_entries, signatures["industryOpportunity"], "industry_opportunity")
        fusion_workbook = identify_fusion_workbook(fusion_entries)

        report = read_json_by_name(extracted["fusion"]["root"], FUSION_REPORT)
        quality = read_json_by_name(extracted["fusion"]["root"], QUALITY_REPORT)
        access = read_json_by_name(extracted["fusion"]["root"], ACCESS_REPORT)
        semantic_rows = table_from_fusion(fusion_workbook, "geo_semantic")
        classified_rows = table_from_fusion(fusion_workbook, "citation_classification_detail")
        expanded_rows = table_from_fusion(fusion_workbook, "geo_keywords_expanded")
        opportunity_top10 = table_from_fusion(fusion_workbook, "opportunity_top10")
        opportunity_all = table_from_fusion(fusion_workbook, "opportunity_all")

        answer_rows = [row for row in ai_answer["rows"] if date_string(row.get("日期")) == target_date]
        citation_rows = [row for row in ai_citation["rows"] if date_string(row.get("日期")) == target_date]
        keyword_rows = [row for row in geo_keyword["rows"] if date_string(row.get("日期")) == target_date]
        answers = build_answer_records(ai_answer["rows"], semantic_rows, target_date, aliases, warnings)
        citations = build_citation_records(ai_citation["rows"], classified_rows, target_date, aliases, warnings)
        keywords = build_keyword_records(expanded_rows, target_date, aliases, warnings)

        platform_answer_counts = collections.Counter(item["platformId"] for item in answers)
        platform_citation_counts = collections.Counter(item["platformId"] for item in citations)
        question_ids = sorted({item["questionId"] for item in answers})
        question_by_id = {item["questionId"]: item["question"] for item in answers}
        platform_ids = sorted(platform_answer_counts)
        brand_roles = collections.Counter(item["diagnostics"]["brandRole"] for item in answers)
        branded_count = sum(1 for item in answers if item["brandMentioned"])
        valid_answer_count = sum(1 for item in answers if item["isValid"])
        invalid_answers = [item for item in answers if not item["isValid"]]
        valid_answers = [item for item in answers if item["isValid"]]
        valid_roles = collections.Counter(item["diagnostics"]["brandRole"] for item in valid_answers)
        validated_brand_mention_count = sum(1 for item in valid_answers if item["brandMentioned"])
        validated_primary_count = valid_roles.get("主推荐", 0)
        validated_secondary_count = valid_roles.get("次推荐", 0)
        validated_brand_recommendation_count = validated_primary_count + validated_secondary_count
        validated_soft_placement_count = valid_roles.get("软植入", 0)
        validated_unmentioned_count = valid_roles.get("未提及", 0)
        metrics_validated = {
            "validAnswerCount": valid_answer_count,
            "brandMentionCount": validated_brand_mention_count,
            "brandMentionRate": round(validated_brand_mention_count / valid_answer_count * 100, 8),
            "primaryRecommendationCount": validated_primary_count,
            "primaryRecommendationRate": round(validated_primary_count / valid_answer_count * 100, 8),
            "secondaryRecommendationCount": validated_secondary_count,
            "secondaryRecommendationRate": round(validated_secondary_count / valid_answer_count * 100, 8),
            "brandRecommendationCount": validated_brand_recommendation_count,
            "brandRecommendationRate": round(validated_brand_recommendation_count / valid_answer_count * 100, 8),
            "softPlacementCount": validated_soft_placement_count,
            "softPlacementRate": round(validated_soft_placement_count / valid_answer_count * 100, 8),
            "unmentionedCount": validated_unmentioned_count,
            "unmentionedRate": round(validated_unmentioned_count / valid_answer_count * 100, 8),
            "denominator": valid_answer_count,
            "ruleBasis": "valid_answers_only",
            "roleDistribution": dict(sorted(valid_roles.items())),
            "roleCountTotal": sum(valid_roles.values()),
            "rolesMutuallyExclusive": sum(valid_roles.values()) == valid_answer_count,
        }
        expected_questions_per_platform = int(business_rules["expectedQuestionCountPerPlatform"])
        actual_combinations = {(item["platformId"], item["questionId"]) for item in answers}
        expected_combinations = {(pid, question_id) for pid in platform_ids for question_id in question_ids}
        missing_combinations = [
            {
                "platformId": pid,
                "questionId": question_id,
                "question": question_by_id[question_id],
                "expected": True,
                "collected": False,
                "reason": "固定18题核心面板在该平台未采集到回答",
            }
            for pid, question_id in sorted(expected_combinations - actual_combinations)
        ]
        missing_by_platform = collections.Counter(item["platformId"] for item in missing_combinations)
        completeness_rate = round(len(actual_combinations) / len(expected_combinations) * 100, 8)
        answer_validity_rate = round(valid_answer_count / len(answers) * 100, 8)
        theoretical_valid_completeness_rate = round(valid_answer_count / len(expected_combinations) * 100, 8)

        source_type_counts = collections.Counter(item["sourceType"] for item in citations)
        content_type_counts = collections.Counter(item["contentType"] for item in citations)
        domain_rows: dict[str, list[dict[str, Any]]] = collections.defaultdict(list)
        for item in citations:
            domain_rows[item["domain"]].append(item)
        source_domains = []
        for domain, items in domain_rows.items():
            types = collections.Counter(item["sourceType"] for item in items)
            source_domains.append({
                "domain": domain,
                "count": len(items),
                "rate": round(len(items) / len(citations) * 100, 8),
                "sourceType": types.most_common(1)[0][0],
                "qualityLevel": None,
                "indexed": None,
                "status": "missing",
            })
        source_domains.sort(key=lambda item: (-item["count"], item["domain"]))

        expected = mapping["expected"]
        reconciliation, check = make_reconciliation(mapping)
        check("AI回答原始累计记录", expected["rawAnswerRows"], len(ai_answer["rows"]), "GEO清洗Excel/AI回答分析表")
        check("AI回答目标日期记录", expected["targetAnswerRows"], len(answer_rows), "GEO清洗Excel/AI回答分析表")
        check("AI回答排除历史记录", expected["excludedAnswerRows"], len(ai_answer["rows"]) - len(answer_rows), "GEO清洗Excel/AI回答分析表")
        check("AI引用记录", expected["rawCitationRows"], len(ai_citation["rows"]), "GEO清洗Excel/AI引用分析表")
        check("AI引用目标日期记录", expected["targetCitationRows"], len(citation_rows), "GEO清洗Excel/AI引用分析表")
        check("GEO来源关键词累计记录", expected["rawKeywordRows"], len(geo_keyword["rows"]), "GEO清洗Excel/GEO关键词研究库")
        check("GEO来源关键词目标日期记录", expected["targetKeywordRows"], len(keyword_rows), "GEO清洗Excel/GEO关键词研究库")
        check("GEO来源关键词排除历史记录", expected["excludedKeywordRows"], len(geo_keyword["rows"]) - len(keyword_rows), "GEO清洗Excel/GEO关键词研究库")
        check("拓展关键词", expected["expandedKeywordRows"], len(expanded_rows), "融合汇总Excel/geo_keywords_expanded")
        check("5A资产记录", expected["a5AssetRows"], len(a5_asset["rows"]), "5A清洗Excel")
        check("5A流转记录", expected["a5FlowRows"], len(a5_flow["rows"]), "5A清洗Excel")
        check("八大人群记录", expected["audienceRows"], len(audience["rows"]), "5A清洗Excel")
        check("品牌心智记录", expected["brandMindRows"], len(brand_mind["rows"]), "品牌心智清洗Excel")
        check("行业机会记录", expected["industryOpportunityRows"], len(industry_opportunity["rows"]), "行业机会清洗Excel")
        check("融合语义连接完整", len(answer_rows), len(semantic_rows), "融合汇总Excel/geo_semantic")
        check("引用分类连接完整", len(citation_rows), len(classified_rows), "融合汇总Excel/citation_classification_detail")
        check("平台回答分布", expected["platformAnswerRows"], dict(sorted(platform_answer_counts.items())), "目标日期回答明细")
        check("平台引用分布", expected["platformCitationRows"], {pid: platform_citation_counts.get(pid, 0) for pid in expected["platformCitationRows"]}, "目标日期引用明细")
        check("引用来源分类", expected["citationSourceTypeCounts"], {key: source_type_counts.get(key, 0) for key in expected["citationSourceTypeCounts"]}, "融合汇总Excel/citation_classification_detail")
        check("理论平台×问题组合", expected["theoreticalPlatformQuestionCombinations"], len(expected_combinations), "规则总表 R-DQ-002")
        check("实际平台×问题组合", expected["actualPlatformQuestionCombinations"], len(actual_combinations), "AI回答明细")
        check("缺失平台×问题组合", expected["missingPlatformQuestionCombinations"], len(missing_combinations), "固定18题面板差集")
        check("有效回答数", expected["validAnswerRows"], valid_answer_count, "AI回答清洗明细+规则变更日志 CHG-028")
        check("无效回答数", expected["invalidAnswerRows"], len(invalid_answers), "AI回答清洗明细+规则变更日志 CHG-028")
        actual_top10 = [row.get("心智词") for row in sorted(opportunity_top10, key=lambda item: number(item.get("execution_priority_rank")) or 999)]
        check("行业机会Top10", expected["industryTop10"], actual_top10, "融合汇总Excel/opportunity_top10")
        check("融合日期", target_date, report.get("data_date"), "融合分析JSON")
        check("5A日期", target_date, report.get("a5_data_date"), "融合分析JSON")
        for key, expected_score in expected["scores"].items():
            check(f"权威评分:{key}", expected_score, report.get("scores", {}).get(key), "融合分析JSON/scores", tolerance=1e-9)

        authoritative_brand_rate = percent(report["geo"]["brand_mention_rate"])
        recomputed_brand_rate = round(branded_count / len(answers) * 100, 8)
        check("品牌提及率逐行核验", authoritative_brand_rate, recomputed_brand_rate, "融合分析JSON vs 回答明细", tolerance=0.01)
        check("无效回答可唯一识别", "answer_equals_question", invalid_answers[0]["diagnostics"]["invalidReason"] if len(invalid_answers) == 1 else None, "AI回答清洗明细源行")

        failed_reconciliation = [item for item in reconciliation if not item["passed"] and item["blockingLevel"] == "error"]
        write_json(diagnostics_dir / "reconciliation.json", {
            "status": "fail" if failed_reconciliation else "pass",
            "checks": reconciliation,
            "blockingFailures": failed_reconciliation,
        })
        if failed_reconciliation:
            raise ConversionError(f"RECONCILIATION_FAILED: {len(failed_reconciliation)} blocking checks")

        display_names = {"all": "全部平台", "doubao": "豆包", "deepseek": "DeepSeek", "kimi": "Kimi", "qwen": "千问"}
        platforms = []
        for pid in ["all", "doubao", "deepseek", "kimi", "qwen"]:
            answer_count = len(answers) if pid == "all" else platform_answer_counts.get(pid, 0)
            citation_count = len(citations) if pid == "all" else platform_citation_counts.get(pid, 0)
            valid_count = valid_answer_count if pid == "all" else sum(1 for item in answers if item["platformId"] == pid and item["isValid"])
            expected_count = len(expected_combinations) if pid == "all" else expected_questions_per_platform
            status = "healthy" if pid == "deepseek" else "warning"
            platforms.append({
                "id": pid,
                "name": display_names[pid],
                "displayName": display_names[pid],
                "aliases": [key for key, value in aliases.items() if value == pid],
                "enabled": True,
                "expectedQuestionCount": expected_count,
                "collectedQuestionCount": answer_count,
                "validAnswerCount": valid_count,
                "accessibilityRate": 100.0 if pid == "all" or answer_count else 0.0,
                "questionCollectionCompleteness": round(answer_count / expected_count * 100, 8),
                "collectedAnswerValidity": round(valid_count / answer_count * 100, 8) if answer_count else None,
                "weight": None,
                "status": status,
                "diagnostics": {
                    "answerRows": answer_count,
                    "citationRows": citation_count,
                    "missingQuestionCount": len(missing_combinations) if pid == "all" else missing_by_platform.get(pid, 0),
                    "citationEvidenceStatus": "unknown_or_no_returned_citation" if pid == "kimi" else "available",
                },
            })

        answer_types = []
        for label, count in sorted(collections.Counter(item["answerType"] for item in valid_answers).items()):
            answer_types.append({"id": stable_id("atype", label, length=12), "label": label, "count": count, "rate": round(count / valid_answer_count * 100, 8)})

        platform_comparison = []
        for pid in ["doubao", "deepseek", "kimi", "qwen"]:
            rows = [item for item in answers if item["platformId"] == pid]
            valid_rows = [item for item in rows if item["isValid"]]
            branded = sum(1 for item in valid_rows if item["brandMentioned"])
            valid = len(valid_rows)
            roles = collections.Counter(item["diagnostics"]["brandRole"] for item in valid_rows)
            platform_comparison.append({
                "platformId": pid,
                "accessibilityRate": 100.0 if rows else 0.0,
                "completenessRate": round(len(rows) / expected_questions_per_platform * 100, 8),
                "validityRate": round(valid / len(rows) * 100, 8) if rows else None,
                "brandMentionRate": round(branded / valid * 100, 8) if valid else None,
                "firstRecommendationRate": None,
                "primaryRecommendationRate": round(roles.get("主推荐", 0) / valid * 100, 8) if valid else None,
                "secondaryRecommendationRate": round(roles.get("次推荐", 0) / valid * 100, 8) if valid else None,
                "brandRecommendationRate": round((roles.get("主推荐", 0) + roles.get("次推荐", 0)) / valid * 100, 8) if valid else None,
                "averageBrandPosition": None,
            })

        top_keywords = []
        for row in sorted(opportunity_top10, key=lambda item: number(item.get("execution_priority_rank")) or 999):
            word = normalize_question(row.get("心智词"))
            top_keywords.append({
                "keywordId": stable_id("opp", target_date, word),
                "keyword": word,
                "normalizedKeyword": word.casefold(),
                "sourceKeyword": word,
                "candidateKeyword": word,
                "platformId": "all",
                "date": target_date,
                "commercialValue": None,
                "aiTriggerType": "unknown",
                "brandOpportunity": number(row.get("brand_match_score")),
                "optimizationDirection": row.get("机会方向") or None,
                "sceneType": row.get("所属维度") or "unknown",
                "trend": "up" if (number(row.get("增长率")) or 0) > 0 else "down" if (number(row.get("增长率")) or 0) < 0 else "flat",
                "trendValue": percent(row.get("增长率")),
                "priority": {"高": "high", "中": "medium", "低": "low"}.get(str(row.get("机会等级") or ""), "unknown"),
                "status": "covered" if row.get("品牌是否覆盖") == "已覆盖" else "opportunity",
                "opportunityScore": number(row.get("opportunity_score")),
                "marketSignalType": row.get("market_signal_type"),
                "opportunityType": row.get("opportunity_type"),
            })

        group_counts = collections.Counter((row.get("opportunity_type") or "unknown", row.get("market_signal_type") or "unknown", row.get("所属维度") or "unknown") for row in opportunity_all)
        opportunity_groups = [
            {"opportunityType": key[0], "marketSignalType": key[1], "dimension": key[2], "count": count}
            for key, count in sorted(group_counts.items(), key=lambda item: (-item[1], item[0]))
        ]
        scene_counts = collections.Counter(item["sceneType"] for item in keywords)

        alerts = [
            {"id": "kimi-citation-status-unknown", "level": "warning", "category": "citation", "title": "Kimi引用状态待确认", "message": "Kimi本批11条回答未返回引用记录，当前无法区分未触发检索、检索无引用或采集状态。", "metricId": "citationEvidenceStatus", "platformId": "kimi", "questionId": None, "date": target_date, "status": "active"},
            {"id": "qwen-answer-sample-low", "level": "warning", "category": "answer", "title": "千问回答样本偏少", "message": "千问AI回答仅3条，低于本批最高平台18条的50%。", "metricId": "answerRecords", "platformId": "qwen", "questionId": None, "date": target_date, "status": "active"},
            {"id": "question-collection-incomplete", "level": "warning", "category": "data-health", "title": "问题采集不完整", "message": "固定18题×4平台共72个组合，实际采集49个，缺失23个。", "metricId": "questionCollectionCompleteness", "platformId": None, "questionId": None, "date": target_date, "status": "active"},
            {"id": "invalid-answer-detected", "level": "warning", "category": "data-health", "title": "存在1条无效回答", "message": "豆包“泡面有什么好吃的”回答内容与问题相同，已从有效回答口径中排除。", "metricId": "collectedAnswerValidity", "platformId": "doubao", "questionId": invalid_answers[0]["questionId"], "date": target_date, "status": "active"},
            {"id": "geo-semantic-score-down", "level": "warning", "category": "trend", "title": "GEO语义评分下降", "message": "2026-07-29 GEO语义评分较2026-07-28下降4.66分。", "metricId": "geoSemanticScore", "platformId": None, "questionId": None, "date": target_date, "status": "active"},
            {"id": "keyword-candidates-untested", "level": "warning", "category": "keyword", "title": "AI拓展候选尚未实测", "message": "39条AI拓展候选未逐条真实提问，测试率与引用触发率按N/A展示。", "metricId": "candidateTestRate", "platformId": None, "questionId": None, "date": target_date, "status": "active"},
        ]
        data_quality_warnings = [
            {"code": "QUESTION_COLLECTION_INCOMPLETE", "message": "固定18题×4平台共72个组合，实际采集49个，完整率68.06%。"},
            {"code": "INVALID_ANSWER_DETECTED", "message": "49条采集回答中有1条回答与问题相同，有效回答48条。"},
            {"code": "KIMI_CITATION_STATUS_UNKNOWN", "message": "Kimi本批11条回答未返回引用记录，当前无法区分未检索、无引用或采集状态。"},
            {"code": "QWEN_SAMPLE_LOW", "message": "千问AI回答仅3条，低于本批最高平台18条的50%。"},
            {"code": "LOW_CONFIDENCE_BATCH", "message": "7月29日批次为低置信观察批次，不进入正式投放趋势。"},
            {"code": "SINGLE_DAY_DETAIL_ONLY", "message": "当前真实明细仅包含单日数据，无法构建相同平台×问题的逐行趋势。"},
        ]
        warnings.extend(data_quality_warnings)

        scores = report["scores"]
        geo = report["geo"]
        combined_input_checksum = sha256_bytes("|".join(sources[key]["sha256"] for key in sorted(sources)).encode("ascii"))
        competitors = split_values(str(report.get("competitor") or "").replace("、", ","))
        competitor_set_version = "competitors-" + sha256_bytes("\n".join(sorted(competitors)).encode("utf-8"))[:12]
        question_set_version = "questions-" + sha256_bytes("\n".join(question_ids).encode("utf-8"))[:12]
        platform_set_version = "platforms-" + sha256_bytes("\n".join(platform_ids).encode("utf-8"))[:12]

        dataset = {
            "schemaVersion": mapping["schemaVersion"],
            "datasetId": args.dataset_id,
            "datasetVersion": mapping["datasetVersion"],
            "source": {
                "type": "json",
                "name": "杨掌柜GEO真实融合数据",
                "fileName": output.name,
                "generatedAt": report["generated_at"],
                "importedAt": None,
                "producer": "ActiveTheory GEO Real Data Converter V1.6.2",
                "sourceVersion": report["report_version"],
                "checksum": combined_input_checksum,
                "notes": ["2026-07-29真实清洗与融合分析数据", "V1历史评分仅供观察", "低置信批次不进入正式趋势", "公开数据包已启用sanitize"],
            },
            "metadata": {
                "reportDate": target_date,
                "geoDataDate": report["date_metadata"]["geo_data_date"],
                "fiveASnapshotDate": business_rules["fiveASnapshotDate"],
                "brandMindSnapshotDate": business_rules["brandMindSnapshotDate"],
                "lagDays": 7,
                "fiveALagDays": 7,
                "brandMindLagDays": 7,
                "timezone": "Asia/Shanghai",
                "dateAlignmentStatus": "warning",
                "competitorSetVersion": competitor_set_version,
                "questionSetVersion": question_set_version,
                "platformSetVersion": platform_set_version,
                "dataWindowStart": target_date,
                "dataWindowEnd": target_date,
                "analysisRuleVersion": business_rules["analysisRuleVersion"],
                "qualityRuleVersion": business_rules["qualityRuleVersion"],
                "dataContractVersion": business_rules["dataContractVersion"],
                "presentationVersion": business_rules["presentationVersion"],
                "confidenceLevel": business_rules["confidenceLevel"],
                "formalTrendEligible": business_rules["formalTrendEligible"],
                "qualityStatus": business_rules["qualityStatus"],
                "ruleDocumentVersion": business_rules["ruleDocumentVersion"],
                "rulesExtractedAt": business_rules["rulesExtractedAt"],
            },
            "platforms": platforms,
            "overview": {
                "finalScore": scores["final_score"],
                "geoStructureScore": scores["geo_structural_quality_score"],
                "geoSemanticScore": scores["geo_semantic_score"],
                "geoScore": scores["geo_score"],
                "fiveAScore": scores["a5_score"],
                "industryOpportunityScore": scores["industry_opportunity_driver_score"],
                "keywordEffectivenessScore": None,
                "brandVisibilityRate": round(percent(geo["brand_mention_rate"]), 2),
                "firstRecommendationRate": None,
                "firstRecommendationDefinitionStatus": "pending_definition",
                "primaryRecommendationRate": round(percent(geo["brand_main_recommend_rate"]), 2),
                "secondaryRecommendationRate": round(percent(geo["brand_secondary_recommend_rate"]), 2),
                "brandRecommendationRate": round(percent(geo["brand_recommend_rate"]), 2),
                "softPlacementRate": round(percent(geo["soft_placement_rate"]), 2),
                "averageBrandPosition": None,
                "qualityCitationRate": None,
                "keywordOpportunityScore": scores["industry_opportunity_driver_score"],
                "dataHealthScore": None,
                "scoreChange": next((number(item["变化"]) for item in report["trend_comparison"]["rows"] if item["指标"] == "综合评分"), -0.23),
                "scoreChangeDirection": "down",
                "scoreComponents": [
                    {"id": "geo-structure", "label": "GEO结构质量", "value": scores["geo_structural_quality_score"], "weight": None, "contribution": None, "sourceMetric": "scores.geo_structural_quality_score", "status": "active"},
                    {"id": "geo-semantic", "label": "GEO语义质量", "value": scores["geo_semantic_score"], "weight": None, "contribution": None, "sourceMetric": "scores.geo_semantic_score", "status": "active"},
                    {"id": "a5", "label": "5A评分", "value": scores["a5_score"], "weight": None, "contribution": None, "sourceMetric": "scores.a5_score", "status": "active"},
                    {"id": "industry", "label": "行业机会驱动", "value": scores["industry_opportunity_driver_score"], "weight": None, "contribution": None, "sourceMetric": "scores.industry_opportunity_driver_score", "status": "active"},
                    {"id": "keyword-effectiveness", "label": "历史关键词有效性", "value": scores["keyword_effectiveness_score"], "weight": None, "contribution": None, "sourceMetric": "scores.keyword_effectiveness_score", "status": "deprecated"},
                ],
            },
            "answer": {
                "summary": {"totalQuestions": len(question_ids), "collectedAnswers": len(answers), "validAnswers": valid_answer_count, "invalidAnswers": len(invalid_answers), "brandedAnswers": branded_count, "firstRecommendations": None},
                "metrics": {"platformAccessibilityRate": 100.0, "questionCollectionCompleteness": completeness_rate, "collectedAnswerValidity": answer_validity_rate, "brandMentionRate": round(percent(geo["brand_mention_rate"]), 2), "firstRecommendationRate": None, "firstRecommendationDefinitionStatus": "pending_definition", "primaryRecommendationRate": round(percent(geo["brand_main_recommend_rate"]), 2), "secondaryRecommendationRate": round(percent(geo["brand_secondary_recommend_rate"]), 2), "brandRecommendationRate": round(percent(geo["brand_recommend_rate"]), 2), "softPlacementRate": round(percent(geo["soft_placement_rate"]), 2), "unmentionedRate": round(percent(geo["unmentioned_rate"]), 2), "averageBrandPosition": None},
                "metricsValidated": metrics_validated,
                "correctedMetrics": {**metrics_validated, "status": "validated_alias"},
                "answerTypes": answer_types,
                "platformComparison": platform_comparison,
                "brandPositions": [],
                "recommendationLevels": {"primary": valid_roles.get("主推荐", 0), "secondary": valid_roles.get("次推荐", 0), "weak": valid_roles.get("多品牌混合/歧义归属", 0) + valid_roles.get("无品牌产品推荐", 0), "softPlacement": valid_roles.get("软植入", 0), "noRecommendation": valid_roles.get("未提及", 0) + valid_roles.get("产品错归竞品", 0)},
                "records": answers,
            },
            "citation": {
                "summary": {"totalCitations": len(citations), "validCitations": len(citations), "qualityCitations": None, "uniqueDomains": len(source_domains)},
                "metrics": {"qualityRate": None, "officialRate": round(source_type_counts.get("official", 0) / len(citations) * 100, 8), "thirdPartyRate": round(source_type_counts.get("thirdParty", 0) / len(citations) * 100, 8), "communityRate": round(source_type_counts.get("community", 0) / len(citations) * 100, 8), "rankingReviewRate": round(content_type_counts.get("review", 0) / len(citations) * 100, 8), "indexedRate": None},
                "sourceTypes": [{"id": key, "count": source_type_counts.get(key, 0), "rate": round(source_type_counts.get(key, 0) / len(citations) * 100, 8)} for key in ("official", "thirdParty", "community", "unknown")],
                "contentTypes": [{"id": key, "count": count, "rate": round(count / len(citations) * 100, 8)} for key, count in sorted(content_type_counts.items())],
                "sourceDomains": source_domains,
                "indexStatus": {"indexed": None, "pending": None, "missing": len(citations), "inaccessible": None},
                "abnormalSources": [{"source": "Kimi", "domain": None, "count": 11, "severity": "medium", "citationEvidenceStatus": "unknown_or_no_returned_citation", "reason": "本批11条回答未返回引用记录，无法区分未检索、无引用或采集状态"}],
                "records": citations,
            },
            "keyword": {
                "summary": {"totalKeywords": len(keywords), "opportunityKeywords": report["geo"]["keyword_analysis"]["industry_cross_match_count"], "sourceKeywordCount": len(keyword_rows), "candidateKeywordCount": len(keywords), "testedCandidateCount": None, "triggeredCandidateCount": None, "opportunityCrossMatchCount": report["geo"]["keyword_analysis"]["industry_cross_match_count"], "newKeywordCount": None, "decliningKeywordCount": None},
                "metrics": {"opportunityScore": scores["industry_opportunity_driver_score"], "industryOpportunityDriverScore": scores["industry_opportunity_driver_score"], "candidateTestRate": None, "candidateCitationTriggerRate": None, "keywordEffectivenessScore": None, "averageCommercialValue": None, "averageBrandOpportunity": None, "highPriorityCount": sum(1 for item in keywords if item["priority"] == "high")},
                "topKeywords": top_keywords,
                "newKeywords": [],
                "decliningKeywords": [],
                "triggerTypes": [{"id": "citationTriggered", "label": "候选引用触发", "count": None, "rate": None, "status": "not_applicable"}, {"id": "untested", "label": "AI拓展候选未实测", "count": len(keywords), "rate": None, "status": "candidate-untested"}],
                "sceneTypes": [{"id": stable_id("scene", key, length=12), "label": key, "count": count} for key, count in sorted(scene_counts.items())],
                "opportunityGroups": opportunity_groups,
                "records": keywords,
            },
            "dataHealth": {
                "platformAccessibility": {"numerator": 4, "denominator": 4, "rate": 100.0, "previousRate": None, "change": None, "status": "healthy", "affectedPlatforms": [], "affectedQuestions": [], "reason": "4个预期平台均有AI回答记录", "recommendation": None},
                "questionCollectionCompleteness": {"numerator": len(actual_combinations), "denominator": len(expected_combinations), "rate": completeness_rate, "previousRate": None, "change": None, "status": "warning", "affectedPlatforms": [{"platformId": pid, "missingCount": missing_by_platform.get(pid, 0)} for pid in platform_ids if missing_by_platform.get(pid, 0)], "affectedQuestions": missing_combinations, "reason": "固定18题×4平台，实际采集49个组合", "recommendation": "补采缺失的23个平台×问题组合"},
                "collectedAnswerValidity": {"numerator": valid_answer_count, "denominator": len(answers), "rate": answer_validity_rate, "previousRate": None, "change": None, "status": "warning", "affectedPlatforms": ["doubao"], "affectedQuestions": [{"questionId": item["questionId"], "question": item["question"], "sourceRow": item["rawReference"]["sourceRow"], "reason": item["diagnostics"]["invalidReason"]} for item in invalid_answers], "reason": "有效回答数 / 实际采集回答数；1条回答与问题相同", "recommendation": "重新采集该问题的有效回答"},
                "theoreticalValidCompleteness": {"numerator": valid_answer_count, "denominator": len(expected_combinations), "rate": theoretical_valid_completeness_rate, "previousRate": None, "change": None, "status": "warning", "affectedPlatforms": ["doubao", "kimi", "qwen"], "affectedQuestions": missing_combinations, "reason": "有效回答数 / 理论平台×问题组合数", "recommendation": "补采缺失组合并替换无效回答"},
                "overallStatus": "warning",
                "qualityStatus": business_rules["qualityStatus"],
                "confidenceLevel": business_rules["confidenceLevel"],
            },
            "trends": [],
            "alerts": alerts,
            "recommendations": [
                {"id": f"execution-{index + 1}", "priority": "high" if index == 0 else "medium", "category": "fusion-execution-order", "title": action, "rationale": "融合分析既有执行顺序", "action": action, "relatedMetric": None, "relatedKeywords": [], "relatedPlatforms": [], "status": "active"}
                for index, action in enumerate(scores["execution_order"])
            ],
            "diagnostics": {
                "growthStage": scores["growth_stage"],
                "bottleneckStage": scores["bottleneck_stage"],
                "topFixPath": scores["top_fix_path"],
                "growthDriver": scores["growth_driver"]["driver_type"],
                "executionOrder": scores["execution_order"],
                "contentSourceQualityScore": geo["content_source_quality"],
                "aggregateTrendComparison": {**report["trend_comparison"], "observationOnly": True, "formalTrend": False, "reason": "样本不完整且平台×问题组合不完全一致"},
                "legacyKeywordMetrics": {"historicalTestedCount": 39, "historicalTriggeredCount": 0, "historicalKeywordEffectivenessScore": scores["keyword_effectiveness_score"], "status": "deprecated", "reason": "候选未逐条真实提问，测试状态失真"},
                "legacyV1AnswerMetrics": {"brandMentionRate": round(percent(geo["brand_mention_rate"]), 2), "primaryRecommendationRate": round(percent(geo["brand_main_recommend_rate"]), 2), "secondaryRecommendationRate": round(percent(geo["brand_secondary_recommend_rate"]), 2), "brandRecommendationRate": round(percent(geo["brand_recommend_rate"]), 2), "softPlacementRate": round(percent(geo["soft_placement_rate"]), 2), "unmentionedRate": round(percent(geo["unmentioned_rate"]), 2), "denominator": len(answers), "status": "historical", "ruleVersion": business_rules["analysisRuleVersion"], "includesInvalidAnswer": True, "formalUse": False},
                "historicalScoreStatus": {"label": "V1历史计算结果", "confidenceLevel": "low", "formalTrendEligible": False, "qualityStatus": "low_confidence"},
                "missingPlatformQuestionCombinations": missing_combinations,
                "invalidAnswerEvidence": [{"platformId": item["platformId"], "questionId": item["questionId"], "question": item["question"], "sourceSheet": item["rawReference"]["sourceSheet"], "sourceRow": item["rawReference"]["sourceRow"], "reason": item["diagnostics"]["invalidReason"], "historicalSemanticCategory": item["diagnostics"]["brandRole"], "answerType": item["answerType"], "brandMentioned": item["brandMentioned"]} for item in invalid_answers],
                "kimiCitationEvidenceStatus": "unknown_or_no_returned_citation",
                "analysisRuleVersion": business_rules["analysisRuleVersion"],
                "qualityRuleVersion": business_rules["qualityRuleVersion"],
                "dataContractVersion": business_rules["dataContractVersion"],
                "presentationVersion": business_rules["presentationVersion"],
                "confidenceLevel": business_rules["confidenceLevel"],
                "formalTrendEligible": business_rules["formalTrendEligible"],
                "qualityStatus": business_rules["qualityStatus"],
                "ruleDocumentVersion": business_rules["ruleDocumentVersion"],
                "rulesExtractedAt": business_rules["rulesExtractedAt"],
                "ruleDocumentSha256": {"ruleTable": rule_sources["ruleTable"]["sha256"], "ruleChangelog": rule_sources["ruleChangelog"]["sha256"]},
                "a5RecordCounts": {"asset": len(a5_asset["rows"]), "flow": len(a5_flow["rows"]), "audience": len(audience["rows"]), "brandMind": len(brand_mind["rows"]), "industryOpportunity": len(industry_opportunity["rows"])},
                "competitors": competitors,
                "warnings": warnings,
                "sourceZipSha256": {key: value["sha256"] for key, value in sources.items()},
                "publicDataSanitized": bool(args.sanitize),
            },
        }

        safety = safety_scan(dataset)
        write_json(diagnostics_dir / "public-data-safety-check.json", safety)
        write_json(correction_dir / "public-data-safety-check-v162.json", safety)
        if safety["status"] != "pass":
            raise ConversionError("PUBLIC_DATA_SAFETY_CHECK_FAILED")

        business_reconciliation = [
            {"ruleId": "R-DQ-002", "ruleName": "问题采集完整率", "previousValue": {"denominator": None, "rate": None}, "correctedValue": {"numerator": len(actual_combinations), "denominator": len(expected_combinations), "rate": completeness_rate, "missingCombinations": len(missing_combinations)}, "evidence": "固定18题×4平台；AI回答明细实际采集49个组合", "sourceDocument": RULE_TABLE_NAME, "status": "corrected", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": False},
            {"ruleId": "R-DQ-003", "ruleName": "已采集回答有效率", "previousValue": {"numerator": 49, "denominator": 49, "rate": 100.0}, "correctedValue": {"numerator": valid_answer_count, "denominator": len(answers), "rate": answer_validity_rate}, "evidence": f"AI回答分析表源行{invalid_answers[0]['rawReference']['sourceRow']}回答与问题完全相同", "sourceDocument": RULE_CHANGELOG_NAME, "status": "corrected", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": False},
            {"ruleId": "R-DATE-003/R-DATE-004", "ruleName": "5A与心智真实快照日期", "previousValue": "2026-07-29", "correctedValue": {"fiveASnapshotDate": business_rules["fiveASnapshotDate"], "brandMindSnapshotDate": business_rules["brandMindSnapshotDate"], "lagDays": 7}, "evidence": "清洗表处理说明显示日期来源为用户指定业务日期；规则总表记录7-22→7-29", "sourceDocument": RULE_TABLE_NAME, "status": "corrected_with_documented_source_conflict", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": False},
            {"ruleId": "R-ANS-021", "ruleName": "首位推荐定义", "previousValue": 0.0, "correctedValue": None, "evidence": "历史资料未给出正式分子、分母和可执行定义", "sourceDocument": RULE_TABLE_NAME, "status": "pending_definition", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": True},
            {"ruleId": "R-ANS-010/R-ANS-011/R-ANS-012", "ruleName": "正式推荐口径", "previousValue": {"firstRecommendationRate": 0.0}, "correctedValue": {"historical": {"primaryRecommendationRate": 0.0, "secondaryRecommendationRate": 8.16, "brandRecommendationRate": 8.16, "softPlacementRate": 4.08, "denominator": 49}, "validated": metrics_validated}, "evidence": "V1融合引擎历史聚合值与48条有效回答逐行语义分别保留", "sourceDocument": RULE_TABLE_NAME, "status": "corrected", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": False},
            {"ruleId": "R-KW-004/R-KW-005/R-KW-006", "ruleName": "候选词测试与有效性", "previousValue": {"testedCount": 39, "triggeredCount": 0, "keywordEffectivenessScore": 30.0}, "correctedValue": {"testedCandidateCount": None, "triggeredCandidateCount": None, "candidateTestRate": None, "candidateCitationTriggerRate": None, "keywordEffectivenessScore": None}, "evidence": "AI拓展候选未逐条真实提问", "sourceDocument": RULE_TABLE_NAME, "status": "deprecated_legacy_preserved", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": False},
            {"ruleId": "R-CIT-011", "ruleName": "Kimi引用证据状态", "previousValue": "citation_detail_missing", "correctedValue": "unknown_or_no_returned_citation", "evidence": "Kimi 11条回答的引用字段实际为[]", "sourceDocument": RULE_CHANGELOG_NAME, "status": "corrected", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": True},
            {"ruleId": "R-DQ-009/R-TREND-005", "ruleName": "低置信与正式趋势资格", "previousValue": {"confidenceLevel": None, "formalTrendEligible": True}, "correctedValue": {"confidenceLevel": "low", "formalTrendEligible": False, "qualityStatus": "low_confidence"}, "evidence": "49/72、千问3/18、1条无效回答且平台样本失衡", "sourceDocument": RULE_CHANGELOG_NAME, "status": "corrected", "confidence": "high", "affectsHistoricalScore": False, "affectsDashboard": True, "requiresFutureConfirmation": False},
        ]
        write_json(correction_dir / "business-rule-reconciliation.json", {"status": "pass", "items": business_reconciliation})
        write_json(correction_dir / "data-quality-correction.json", {"theoreticalCombinations": len(expected_combinations), "actualCombinations": len(actual_combinations), "completenessRate": completeness_rate, "missingCombinationCount": len(missing_combinations), "missingCombinations": missing_combinations, "collectedAnswers": len(answers), "validAnswers": valid_answer_count, "invalidAnswers": len(invalid_answers), "answerValidityRate": answer_validity_rate, "theoreticalValidCompletenessRate": theoretical_valid_completeness_rate, "invalidAnswerEvidence": dataset["diagnostics"]["invalidAnswerEvidence"]})
        write_json(correction_dir / "snapshot-date-correction.json", {"reportDate": target_date, "geoDataDate": target_date, "fiveASnapshotDate": business_rules["fiveASnapshotDate"], "brandMindSnapshotDate": business_rules["brandMindSnapshotDate"], "fiveALagDays": 7, "brandMindLagDays": 7, "cleanedDateConflict": {"cleanedTablesDate": report["date_metadata"]["a5_data_date"], "dateSource": "用户指定业务日期", "status": "historical_v1_date_rewrite"}})
        write_json(correction_dir / "recommendation-metric-correction.json", {"firstRecommendationRate": None, "firstRecommendationDefinitionStatus": "pending_definition", "legacyV1AnswerMetrics": dataset["diagnostics"]["legacyV1AnswerMetrics"], "metricsValidated": dataset["answer"]["metricsValidated"], "invalidAnswerHistoricalSemanticCategory": invalid_answers[0]["diagnostics"]["brandRole"], "firstCharacterPositionMeaning": "text_character_position_not_recommendation_rank"})
        write_json(correction_dir / "keyword-metric-correction.json", {"current": {"sourceKeywordCount": len(keyword_rows), "candidateKeywordCount": len(keywords), "testedCandidateCount": None, "triggeredCandidateCount": None, "candidateTestRate": None, "candidateCitationTriggerRate": None, "keywordEffectivenessScore": None, "industryOpportunityDriverScore": scores["industry_opportunity_driver_score"], "opportunityCrossMatchCount": report["geo"]["keyword_analysis"]["industry_cross_match_count"]}, "legacy": dataset["diagnostics"]["legacyKeywordMetrics"]})
        write_json(correction_dir / "kimi-citation-status-correction.json", {"answerCount": platform_answer_counts.get("kimi", 0), "citationRecordCount": platform_citation_counts.get("kimi", 0), "citationEvidenceStatus": "unknown_or_no_returned_citation", "message": "Kimi本批11条回答未返回引用记录，当前无法区分未触发检索、检索无引用或采集状态。", "captureFailureAsserted": False})
        write_json(correction_dir / "version-metadata.json", {key: business_rules[key] for key in ("analysisRuleVersion", "qualityRuleVersion", "dataContractVersion", "presentationVersion", "confidenceLevel", "formalTrendEligible", "qualityStatus", "ruleDocumentVersion", "rulesExtractedAt")})

        validation = {
            "status": "warning",
            "schemaVersion": dataset["schemaVersion"],
            "datasetId": dataset["datasetId"],
            "coreModulesPresent": all(key in dataset for key in ("metadata", "platforms", "overview", "answer", "citation", "keyword", "dataHealth", "trends", "alerts", "recommendations")),
            "warnings": data_quality_warnings,
            "errors": [],
            "sourceQualityStatus": quality.get("status"),
            "sourceAccessStatus": access.get("status"),
        }
        write_json(diagnostics_dir / "real-data-validation.json", validation)

        conversion_diagnostics = {
            "converter": "ActiveTheory GEO Real Data Converter V1.6.2",
            "pythonVersion": os.sys.version.split()[0],
            "openpyxlVersion": openpyxl.__version__,
            "targetDate": target_date,
            "datasetId": args.dataset_id,
            "inputMembers": {key: value["members"] for key, value in extracted.items()},
            "identifiedTables": {
                "aiAnswer": {"sheet": ai_answer["sheet"], "headers": ai_answer["headers"], "rows": len(ai_answer["rows"])},
                "aiCitation": {"sheet": ai_citation["sheet"], "headers": ai_citation["headers"], "rows": len(ai_citation["rows"])},
                "geoKeyword": {"sheet": geo_keyword["sheet"], "headers": geo_keyword["headers"], "rows": len(geo_keyword["rows"])},
                "a5Asset": {"sheet": a5_asset["sheet"], "rows": len(a5_asset["rows"])},
                "a5Flow": {"sheet": a5_flow["sheet"], "rows": len(a5_flow["rows"])},
                "audience": {"sheet": audience["sheet"], "rows": len(audience["rows"])},
                "brandMind": {"sheet": brand_mind["sheet"], "rows": len(brand_mind["rows"])},
                "industryOpportunity": {"sheet": industry_opportunity["sheet"], "rows": len(industry_opportunity["rows"])},
            },
            "joins": {"answerSemanticRows": len(semantic_rows), "citationClassificationRows": len(classified_rows), "duplicateAnswerIds": len(answers) - len({item["recordId"] for item in answers}), "duplicateCitationIds": len(citations) - len({item["citationId"] for item in citations})},
            "warnings": warnings,
            "temporaryExtractionCleanedOnExit": True,
        }
        write_json(diagnostics_dir / "conversion-diagnostics.json", conversion_diagnostics)
        (diagnostics_dir / "field-mapping.md").write_text(field_mapping_markdown(), encoding="utf-8", newline="\n")
        correction_report = f"""# GEO Dashboard V1.6.2 Business Rule Correction

## 规则来源

- `{rule_sources['ruleTable']['path']}`
  SHA-256: `{rule_sources['ruleTable']['sha256']}`
- `{rule_sources['ruleChangelog']['path']}`
  SHA-256: `{rule_sources['ruleChangelog']['sha256']}`
- 使用时间：{rule_sources['usedAt']}
- 规则版本：{business_rules['ruleDocumentVersion']}
- 抽取统计：有效 {rule_sources['ruleCounts']['effective']}，已废弃 {rule_sources['ruleCounts']['deprecated']}，待确认 {rule_sources['ruleCounts']['pendingConfirmation']}，总计 {rule_sources['ruleCounts']['total']}。

## 核心校正

- 问题完整率：{len(actual_combinations)}/{len(expected_combinations)} = {completeness_rate:.2f}%，缺失 {len(missing_combinations)} 个平台×问题组合。
- 回答有效率：{valid_answer_count}/{len(answers)} = {answer_validity_rate:.2f}%；无效证据为豆包源行 {invalid_answers[0]['rawReference']['sourceRow']} 回答与问题相同。
- 理论有效完整率：{valid_answer_count}/{len(expected_combinations)} = {theoretical_valid_completeness_rate:.2f}%。
- 5A与品牌心智真实快照日期恢复为 2026-07-22，lag=7 天；清洗表中的 2026-07-29 记录为 V1 用户指定业务日期改写。
- 首位推荐保持 N/A/待定义；主推荐 0.00%，次推荐 8.16%，品牌推荐 8.16%，软植入 4.08%。
- 无效回答在V1逐行语义中属于“{invalid_answers[0]['diagnostics']['brandRole']}”；有效回答口径为品牌提及 {metrics_validated['brandMentionCount']}/{valid_answer_count}、主推荐 {metrics_validated['primaryRecommendationCount']}/{valid_answer_count}、次推荐 {metrics_validated['secondaryRecommendationCount']}/{valid_answer_count}、软植入 {metrics_validated['softPlacementCount']}/{valid_answer_count}、未提及 {metrics_validated['unmentionedCount']}/{valid_answer_count}。
- 39 条 AI 拓展候选未逐条实测；历史“已测试39/有效性30”仅保留在 legacy diagnostics。
- Kimi 引用状态为 `unknown_or_no_returned_citation`，未判定为抓取失败。
- 该批次为低置信观察批次，`formalTrendEligible=false`。V1 历史评分保留，不重算。

## 安全与结论

- 公开数据安全扫描：{safety['status']}
- Data Gate 预期：warning，fallback=false
- 本轮未修改源 ZIP、历史评分或 Dashboard 视觉主体。
"""
        (correction_dir / "geo-dashboard-v162-report.md").write_text(correction_report, encoding="utf-8", newline="\n")

        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_bytes(json_bytes(dataset))
        output_hash = sha256_file(output)
        output_size = output.stat().st_size

        manifest_path = Path(getattr(args, "manifest", None) or MANIFEST_PATH).resolve()
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        entry = {
            "id": args.dataset_id,
            "label": "杨掌柜 GEO真实数据 2026-07-29",
            "file": output.name,
            "schemaVersion": mapping["schemaVersion"],
            "datasetVersion": mapping["datasetVersion"],
            "sha256": output_hash,
            "sizeBytes": output_size,
            "enabled": True,
            "notes": "2026-07-29真实清洗与融合分析数据；49/72组合、48/49有效回答；低置信观察批次，不进入正式趋势",
        }
        datasets = [item for item in manifest.get("datasets", []) if item.get("id") != args.dataset_id]
        datasets.append(entry)
        manifest["datasets"] = datasets
        manifest["generatedAt"] = report["generated_at"]
        manifest_path.write_bytes(json_bytes(manifest))

        result = {
            "status": "success",
            "datasetId": args.dataset_id,
            "output": str(output),
            "sizeBytes": output_size,
            "sha256": output_hash,
            "gate": "warning",
            "counts": {"answers": len(answers), "citations": len(citations), "sourceKeywords": len(keyword_rows), "expandedKeywords": len(keywords)},
            "warningCount": len(data_quality_warnings),
            "diagnosticsDir": str(diagnostics_dir),
        }
        return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--geo-zip", required=True)
    parser.add_argument("--a5-zip", required=True)
    parser.add_argument("--fusion-zip", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--diagnostics-dir", required=True)
    parser.add_argument("--target-date", required=True)
    parser.add_argument("--dataset-id", required=True)
    parser.add_argument("--rule-table", required=True, help=f"Path to {RULE_TABLE_NAME}")
    parser.add_argument("--rule-changelog", required=True, help=f"Path to {RULE_CHANGELOG_NAME}")
    parser.add_argument("--mapping", default=str(DEFAULT_MAPPING))
    parser.add_argument("--manifest", default=str(MANIFEST_PATH), help="Manifest path; primarily useful for isolated converter verification.")
    parser.add_argument("--sanitize", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = convert(args)
    except ConversionError as error:
        print(json.dumps({"status": "failed", "error": str(error)}, ensure_ascii=False, indent=2))
        return 2
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
